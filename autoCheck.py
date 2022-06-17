import openpyxl as xl
from openpyxl.styles import Alignment, Font
from openpyxl.utils.exceptions import InvalidFileException
import os

leftAlign = Alignment(horizontal='left')
rightAlign = Alignment(horizontal='right')

class Good:

    def __init__(self, name) -> None:
        self.name = name
        self.mailIndex = 0x1F1F1F
        self.saleDict = {}
        self.mailDict = {}
        self.saleIdDict = {}    #存原数调单号
        self.mailIdDict = {}    #存补邮数调单号

    def addRecord(self, name, sale, saleId):
        saleRecord = {
            "id":int(saleId),
            "sale":int(sale)
        }
        self.saleDict[name] = saleRecord
        # self.saleDict[name] = sale
        # self.saleIdDict[name] = saleId

    def addMailRecord(self, name, sale, mailId, des):
        mailRecord = {
            "id":int(mailId),
            "sale":int(sale),
            "des":des
        }
        self.mailDict[name] = mailRecord
        # self.mailDict[name] = sale
        # self.mailIdDict[name] = mailId

    def print(self):
        print(self.name)
        for key in self.saleDict:
            print("{" + str(self.saleDict[key]["id"]) + "\t"+ key + "\t"+ str(self.saleDict[key]["sale"]) + "}")
        for key in self.mailDict:
            print("{" + str(self.mailDict[key]["id"]) + "\t"+ key + "\t"+ str(self.mailDict[key]["sale"]) + "\t"+ str(self.mailDict[key]["des"]) + "}")
        # print(self.saleDict)
        # print(self.mailDict)
    
    def getName(self):
        return self.name
    
    def getSale(self):
        return self.saleDict

    def getSaleRecordId(self, name):
        if name in self.saleDict:
            return self.saleDict[name]["id"]
        else:
            return None

    def getMail(self):
        return self.mailDict

    def checkSaleCustomer(self, customer):
        if customer in self.saleDict:
            return True
        # if customer in self.mailDict:
        #     return True
        return False

def buildSale(table):
    row_max = table.max_row
    col_count = table.max_column - 14
    col_offset = 14 # offset所在的列是第一个商品的列
    goodList = []
    # names = []
    
    # 登记所有商品名称
    for col in table.iter_cols(min_col=15,values_only=True):
        good = Good(col[2])
        goodList.append(good)

    for row in table.iter_rows(min_row=4, max_row=row_max, values_only=True):
        if row[1] == None:
            break
        # names.append(row[1])
        name = row[1]
        number = row[0]
        for index in range(col_offset, col_offset + col_count):
            if row[index] == None:
                continue
            good = goodList[index - col_offset]
            # 给重名的添加哈希尾缀，哈希值为名字第一位的ASCII码 * 7 % 10，若还有重名则再加名字第一位的ASCII码 * 7
            oname = name
            temp = 0
            while True:
                if not good.checkSaleCustomer(name):
                    break
                temp = int(ord(oname[0])) * 7 + temp
                name = oname + " (重名#" + str(temp % 10) + " 或补邮备注填错)"
            good.addRecord(name, row[index], number)


    # for good in goodList:
    #     good.print()

    # for col in table.iter_cols(min_col=15,values_only=True):
    #     good = Good(col[2])
    #     for i in range(len(names)):
    #         if col[i+3] == None:
    #             continue
    #         # print(names[i], col[i+3])

    #         # 给重名的添加哈希尾缀，哈希值为名字第一位的ASCII码 * 7 % 10，若还有重名则再加名字第一位的ASCII码 * 7
    #         name = names[i]
    #         temp = 0
    #         while True:
    #             if not good.checkCustomer(names[i]):
    #                 break
    #             temp = int(ord(name[0])) * 7 + temp
    #             names[i] = name + " (重名#" + str(temp % 10) + ")"
    #         good.addRecord(names[i], col[i+3], table.cell(row=4, column=0))
    #     goodList.append(good)
    return goodList

def buildMail(table, goodList):
    row_max = table.max_row
    names = []

    for row in table.iter_rows(min_row=4, max_row=row_max, values_only=True):
        if row[1] == None:
            break
        names.append(row[1])

    for col in table.iter_cols(min_col=15,values_only=True):
        index = -1
        for i in range(len(goodList)):
            if goodList[i].getName() ==  col[2]:
                index = i
        if index == -1:
            continue
        for i in range(len(names)):
            if col[i+3] == None:
                continue
            # print(names[i], col[i+3])
            # 给重名的添加哈希尾缀，哈希值为名字第一位的ASCII码 * 7 % 10，若还有重名则再加名字第一位的ASCII码 * 7
            good = goodList[index]
            mailId = table.cell(row = i+4, column = 1).value
            des = (table.cell(row = i+4, column = 3)).value
            name = names[i]
            temp = 0
            # 为了尽量不让备注填错的人被认为是不同的人，设定在第一次哈希重命名之后先看看新名字有没有在原数调名单中，要是在，则找到备注中含有原数调
            # 单号的补邮数调为止，要是新名字不在（没有重名的情况下），则不管备注中是否含有原数调单号，都认为是一个人
            flag = 0
            while True:
                if not good.checkSaleCustomer(names[i]):
                    break
                if str(good.getSaleRecordId(names[i])) in str(des):
                    break
                temp = int(ord(name[0])) * 7 + temp
                names[i] = name + " (重名#" + str(temp % 10) + " 或补邮备注填错)"
                if flag == 0 and (not good.checkSaleCustomer(names[i])):
                    names[i] = name
                    break
                flag = 1
            goodList[index].addMailRecord(names[i], col[i+3], mailId, des)

    # for good in goodList:
    #     good.print()
    return goodList

def generateRes(goodList):
    workbook = xl.Workbook()
    desLen = -1

    errorTable = workbook["Sheet"]
    errorTable.title = '错误汇总'
    errorTable.column_dimensions['A'].width = 20
    errorTable.cell(row = 1, column = 2).value = "原单号"
    errorTable.cell(row = 1, column = 3).value = "补邮单号"
    errorTable.cell(row = 1, column = 4).value = "id"
    errorTable.column_dimensions['D'].width = 17
    errorTable.cell(row = 1, column = 5).value = "原数量"
    errorTable.cell(row = 1, column = 6).value = "补邮数量"

    pointer = 1
    lastPointer = 1

    for good in goodList:
        desLen = -1
        table = workbook.create_sheet(title = good.getName())
        saleDict = good.getSale()
        mailDict = good.getMail()
        names1 = list(saleDict.keys())
        names2 = list(mailDict.keys())
        names = list(set(names1).union(set(names2)))
        table.cell(row = 1, column = 1).value = "原单号"
        table.cell(row = 1, column = 2).value = "补邮单号"
        table.cell(row = 1, column = 3).value = "id"
        table.cell(row = 1, column = 4).value = "原数量"
        table.cell(row = 1, column = 5).value = "补邮数量"
        table.cell(row = 1, column = 6).value = "小糊涂蛋check"
        table.cell(row = 1, column = 7).value = "补邮备注"
        for i in range(len(names)):
            table.cell(row = i+2, column = 3).value = names[i]
            if names[i] in saleDict:
                table.cell(row = i+2, column = 1).value = saleDict[names[i]]["id"]
                table.cell(row = i+2, column = 4).value = saleDict[names[i]]["sale"]
            else:
                table.cell(row = i+2, column = 1).value = "-"
                table.cell(row = i+2, column = 1).alignment = rightAlign
                table.cell(row = i+2, column = 4).value = "无数据"
                table.cell(row = i+2, column = 4).alignment = rightAlign
            if names[i] in mailDict:
                table.cell(row = i+2, column = 2).value = mailDict[names[i]]["id"]
                table.cell(row = i+2, column = 5).value = mailDict[names[i]]["sale"]
                table.cell(row = i+2, column = 7).value = str(mailDict[names[i]]["des"]) if mailDict[names[i]]["des"] != None else None
                if desLen < len(str(mailDict[names[i]]["des"])):
                    desLen = len(str(mailDict[names[i]]["des"]))
            else:
                table.cell(row = i+2, column = 2).value = "-"
                table.cell(row = i+2, column = 2).alignment = rightAlign
                table.cell(row = i+2, column = 5).value = "无数据"
                table.cell(row = i+2, column = 5).alignment = rightAlign
                table.cell(row = i+2, column = 7).value = None
            if table.cell(row = i+2, column = 4).value != table.cell(row = i+2, column = 5).value:
                table.cell(row = i+2, column = 6).value = "←小糊涂蛋"
                pointer += 1
                if pointer == lastPointer + 1:
                    errorTable.cell(row = pointer, column = 1).value = good.getName()
                errorTable.cell(row = pointer, column = 2).value = table.cell(row = i+2, column = 1).value
                errorTable.cell(row = pointer, column = 2).alignment = rightAlign
                errorTable.cell(row = pointer, column = 3).value = table.cell(row = i+2, column = 2).value
                errorTable.cell(row = pointer, column = 3).alignment = rightAlign
                errorTable.cell(row = pointer, column = 4).value = table.cell(row = i+2, column = 3).value
                errorTable.cell(row = pointer, column = 5).value = table.cell(row = i+2, column = 4).value
                errorTable.cell(row = pointer, column = 5).alignment = rightAlign
                errorTable.cell(row = pointer, column = 6).value = table.cell(row = i+2, column = 5).value
                errorTable.cell(row = pointer, column = 6).alignment = rightAlign

            else:
                table.cell(row = i+2, column = 6).value = None

        table.column_dimensions['G'].width = desLen + 5 if desLen > 14.75 else 14.75
        table.column_dimensions['F'].width = 15
        table.column_dimensions['C'].width = 17

        if pointer > lastPointer + 1:
            errorTable.merge_cells("A"+str(lastPointer + 1)+":A"+str(pointer))

        lastPointer = pointer

    align_center = Alignment(horizontal='center', vertical='center')
    font_bold = Font(bold=True)

    area = errorTable['A']
    for i in area:
        i.alignment = align_center
        i.font = font_bold

    workbook.save('result.xlsx')

def readFile():
    f = open("config.txt", "r")
    mail = f.readline()[:-1]
    goodstr = f.readline()[:-1]

    f.close()
    return [mail, goodstr]

def saveConfig(mail, good):
    if os.path.exists("config.txt"):
        choice = int(input("\n是否用本次输入的文件名数据覆盖原有的config.txt，以便下次直接使用新数据？（1：是；2：否）\n"))
    else:
        choice = int(input("\n是否将本次输入的文件名保存为config.txt，以便下次直接使用？（1：是；2：否）\n"))

    if choice != 1:
        print("已取消\n")
        return
    
    with open("config.txt", "w+") as f:
        f.write(mail+'\n')
        f.write(good+'\n\n\n')
        f.write("数据格式说明："+'\n')
        f.write("第一行：补邮表格文件名（仅支持xlsx后缀名文件）"+'\n')
        f.write("第二行：商品表格文件名（仅支持xlsx后缀名文件），中间用英文逗号隔开"+'\n')
        f.write("可以在符合数据格式的情况下，直接修改该文件"+'\n')

    print("已生成\n")

def isOpen(filename):
    try:
        f = open(filename, "w")
        f.close()
        return False
    except Exception as e:
        if "[Errno 13] Permission denied" in str(e):
            return True
        else:
            return False
    

if __name__ == "__main__":
    print("使用说明：\n")
    print("使用时需要将【此程序放在补邮和商品导出的表格的同一文件夹内】，你要用到什么文件，就把这个文件放在这个程序所在的文件夹里\n")
    print("导出补邮及商品数调表格时，请保证【补邮数调和原数调内同一商品的名字相同】，否则将无法识别！！！\n")
    print("工作完成后，将会输出 result.xlsx 结果表格，若需要重新生成，只需要重新运行程序即可（result.xlsx需要保持关闭状态）\n")
    print("若result.xlsx中出现重名提示，请谨慎判断是重名还是补邮数调备注填写错误\n")
    print("---------------------------------------------------------------------------------------------来自贴心的芝麻丸披萨\n\n\n")

    while isOpen("result.xlsx"):
        input("检测到\"result.xlsx\"文件存在，请关闭该文件后回车")

    mode = -1

    if os.path.exists("config.txt"):
        mode = int(input("检测到config文件存在，是否直接使用该文件的参数？（1：是；2：否）\n"))
        if mode == 1:
            results = readFile()
            mailFilename = results[0]
            goodstr = results[1]

    if mode != 1:
        mailFilename = input('请输入补邮表格名称（以xlsx为后缀名，例如：补邮.xlsx）：\n')
        goodstr = input('请输入商品表格名称（以xlsx为后缀名，并用【英文】逗号隔开，例如：商品1.xlsx,商品2.xlsx）：\n')
        
    goodsFilenames = [x.strip() for x in goodstr.split(',')]
    print("正在工作，请稍候……")
    goodList = []

    try:

        mailTable = xl.load_workbook(mailFilename)["表2"]
        for goodFilename in goodsFilenames:
            goodTable = xl.load_workbook(goodFilename)["表2"]
            goodList.extend(buildSale(goodTable))
        goodList = buildMail(mailTable, goodList)

        # for item in goodList:
        #     item.toString()

        generateRes(goodList)

        print("\n结果表格\"result.xlsx\"已生成，请查看程序所在文件夹\n")

        if mode != 1:
            saveConfig(mailFilename, goodstr)

    except  FileNotFoundError as e:
        print(e)
        print("该名称的文件不存在，请检查：")
        print("\t1. 名称拼写错误")
        print("\t2. 输入多个文件名时未用英文逗号隔开")
        print("\t3. 未将文件放到与本程序同一个文件夹内（可以都放在桌面上）")

    except InvalidFileException as e:
        print(e)
        print("请使用xlsx文件")


    except Exception as e:
        print(e)
        print("未知错误，请联系")
        

    finally:
        input("输入任意键结束")




